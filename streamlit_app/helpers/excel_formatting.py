from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.worksheet import Worksheet

def apply_base_formatting(worksheet: Worksheet) -> None:
    """Apply base formatting to a worksheet including:
    - Remove gridlines
    - Set default font
    - Other base styling as needed
    """
    worksheet.sheet_view.showGridLines = False

def format_header(worksheet: Worksheet, row: int, start_col: int, end_col: int, text: str) -> None:
    """Format a merged header cell with standard styling
    
    Args:
        worksheet: The worksheet to format
        row: The row number for the header
        start_col: Starting column number for merge
        end_col: Ending column number for merge
        text: Header text
    """
    # Add and merge header
    worksheet.cell(row=row, column=start_col, value=text)
    worksheet.merge_cells(
        start_row=row,
        start_column=start_col,
        end_row=row,
        end_column=end_col
    )
    
    # Apply formatting
    cell = worksheet.cell(row=row, column=start_col)
    cell.font = Font(bold=True)
    cell.alignment = Alignment(horizontal='center') 