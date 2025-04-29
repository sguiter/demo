from datetime import date, datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def formatting_page(ws, max_row=None, max_col=None):
    """
    Apply default formatting to the worksheet:
      - hide ALL gridlines (view and print)
      - apply a clean Calibri font, white fill, centered alignment

    Parameters:
    ws (Worksheet): The openpyxl worksheet object.
    max_row (int, optional): Maximum row to process (defaults to ws.max_row).
    max_col (int, optional): Maximum column to process (defaults to ws.max_column).
    """
    # determine processing bounds
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column

    # hide gridlines in the sheet view
    ws.sheet_view.showGridLines = False
    # hide gridlines when printing
    ws.print_options.gridLines = False

    # define styles (no borders)
    default_font      = Font(name='Calibri', size=11)
    default_fill      = PatternFill(fill_type='solid', start_color='FFFFFF')
    default_alignment = Alignment(horizontal='center', vertical='center')

    # apply styles + date formatting
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font      = default_font
            cell.fill      = default_fill
            cell.alignment = default_alignment

            if cell.value is None:
                continue
            

ACRONYM_MAP = {
    # Income Statement
    'revenue': 'Revenue',
    'costOfRevenue': 'Cost of Revenue',
    'grossProfit': 'Gross Profit',
    'grossProfitRatio': 'Gross Profit %',
    'researchAndDevelopmentExpenses': 'R&D Expense',
    'generalAndAdministrativeExpenses': 'G&A Expense',
    'sellingAndMarketingExpenses': 'S&M Expense',
    'sellingGeneralAndAdministrativeExpenses': 'SG&A Expense',
    'otherExpenses': 'Other Expenses',
    'operatingExpenses': 'Operating Expenses',
    'costAndExpenses': 'Cost & Expenses',
    'interestIncome': 'Interest Income',
    'interestExpense': 'Interest Expense',
    'depreciationAndAmortization': 'Depreciation & Amortization',
    'ebitda': 'EBITDA',
    'ebitdaratio': 'EBITDA %',
    'operatingIncome': 'Operating Income',
    'operatingIncomeRatio': 'Operating Income %',
    'totalOtherIncomeExpensesNet': 'Total Other Income & Expenses',
    'incomeBeforeTax': 'Income Before Tax',
    'incomeBeforeTaxRatio': 'Income Before Tax %',
    'incomeTaxExpense': 'Income Tax Expense',
    'netIncome': 'Net Income',
    'netIncomeRatio': 'Net Income %',
    'eps': 'EPS',
    'epsdiluted': 'EPS Diluted',
    'weightedAverageShsOut': 'Wtd Avg Shares Outstanding',
    'weightedAverageShsOutDil': 'Wtd Avg Shares Outstanding Diluted',
    # Balance Sheet
    'cashAndCashEquivalents': 'Cash & Cash Equivalents',
    'shortTermInvestments': 'Short-Term Investments',
    'cashAndShortTermInvestments': 'Cash & Short-Term Investments',
    'netReceivables': 'Net Receivables',
    'inventory': 'Inventory',
    'otherCurrentAssets': 'Other Current Assets',
    'totalCurrentAssets': 'Total Current Assets',
    'propertyPlantEquipmentNet': 'PP&E, Net',
    'goodwill': 'Goodwill',
    'intangibleAssets': 'Intangible Assets',
    'goodwillAndIntangibleAssets': 'Goodwill & Intangible Assets',
    'longTermInvestments': 'Long-Term Investments',
    'taxAssets': 'Tax Assets',
    'otherNonCurrentAssets': 'Other Non-Current Assets',
    'totalNonCurrentAssets': 'Total Non-Current Assets',
    'otherAssets': 'Other Assets',
    'totalAssets': 'Total Assets',
    'accountPayables': 'Accounts Payable',
    'shortTermDebt': 'Short-Term Debt',
    'taxPayables': 'Tax Payables',
    'deferredRevenue': 'Deferred Revenue',
    'otherCurrentLiabilities': 'Other Current Liabilities',
    'totalCurrentLiabilities': 'Total Current Liabilities',
    'longTermDebt': 'Long-Term Debt',
    'deferredRevenueNonCurrent': 'Deferred Revenue (Non-Current)',
    'deferredTaxLiabilitiesNonCurrent': 'Deferred Tax Liabilities (Non-Current)',
    'otherNonCurrentLiabilities': 'Other Non-Current Liabilities',
    'totalNonCurrentLiabilities': 'Total Non-Current Liabilities',
    'otherLiabilities': 'Other Liabilities',
    'capitalLeaseObligations': 'Capital Lease Obligations',
    'totalLiabilities': 'Total Liabilities',
    'preferredStock': 'Preferred Stock',
    'commonStock': 'Common Stock',
    'retainedEarnings': 'Retained Earnings',
    'accumulatedOtherComprehensiveIncomeLoss': 'Accumulated Other Comprehensive Income (Loss)',
    'othertotalStockholdersEquity': "Other Total Stockholders' Equity",
    'totalStockholdersEquity': "Total Stockholders' Equity",
    'totalEquity': 'Total Equity',
    'totalLiabilitiesAndStockholdersEquity': 'Total Liabilities & Stockholders Equity',
    'minorityInterest': 'Minority Interest',
    'totalLiabilitiesAndTotalEquity': 'Total Liabilities & Total Equity',
    'totalInvestments': 'Total Investments',
    'totalDebt': 'Total Debt',
    'netDebt': 'Net Debt',
    # Cash Flow
    'deferredIncomeTax': 'Deferred Income Tax',
    'stockBasedCompensation': 'SBC',
    'changeInWorkingCapital': 'Change in Working Capital',
    'accountsReceivables': 'Accounts Receivable',
    'accountsPayables': 'Accounts Payable',
    'otherWorkingCapital': 'Other Working Capital',
    'otherNonCashItems': 'Other Non-Cash Items',
    'netCashProvidedByOperatingActivities': 'Cash Flow from Operations (CFO)',
    'investmentsInPropertyPlantAndEquipment': 'Investments in PP&E',
    'acquisitionsNet': 'Acquisitions, Net',
    'purchasesOfInvestments': 'Purchases of Investments',
    'salesMaturitiesOfInvestments': 'Sales/Maturities of Investments',
    'otherInvestingActivites': 'Other Investing Activities',
    'netCashUsedForInvestingActivites': 'Cash Flow Used in Investing (CFI)',
    'debtRepayment': 'Debt Repayment',
    'commonStockIssued': 'Common Stock Issued',
    'commonStockRepurchased': 'Common Stock Repurchased',
    'dividendsPaid': 'Dividends Paid',
    'otherFinancingActivites': 'Other Financing Activities',
    'netCashUsedProvidedByFinancingActivities': 'Cash Flow from Financing (CFF)',
    'effectOfForexChangesOnCash': 'Effect of Forex Changes on Cash',
    'netChangeInCash': 'Net Change in Cash',
    'cashAtEndOfPeriod': 'Cash at End of Period',
    'cashAtBeginningOfPeriod': 'Cash at Beginning of Period',
    'operatingCashFlow': 'Operating Cash Flow',
    'capitalExpenditure': 'CapEx',
    'freeCashFlow': 'Free Cash Flow',
}

def get_display_name(key):
    return ACRONYM_MAP.get(key, key)

def formatting_page(ws, max_row=None, max_col=None):
    """
    Apply default formatting to the worksheet:
      - hide ALL gridlines (view and print)
      - apply a clean Calibri font, white fill, centered alignment

    Parameters:
    ws (Worksheet): The openpyxl worksheet object.
    max_row (int, optional): Maximum row to process (defaults to ws.max_row).
    max_col (int, optional): Maximum column to process (defaults to ws.max_column).
    """
    # determine processing bounds
    if max_row is None:
        max_row = ws.max_row
    if max_col is None:
        max_col = ws.max_column

    # hide gridlines in the sheet view
    ws.sheet_view.showGridLines = False
    # hide gridlines when printing
    ws.print_options.gridLines = False

    # define styles (no borders)
    default_font      = Font(name='Calibri', size=11)
    default_fill      = PatternFill(fill_type='solid', start_color='FFFFFF')
    default_alignment = Alignment(horizontal='center', vertical='center')

    # apply styles + date formatting
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col):
        for cell in row:
            cell.font      = default_font
            cell.fill      = default_fill
            cell.alignment = default_alignment

            if cell.value is None:
                continue
            