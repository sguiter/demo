import os
import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from streamlit_app.helpers.API_helpers import get_financial_statement

# ────────────────────────────────────────────────────────────────────────────────
# External helper – optional FMP call for shares-outstanding
# ────────────────────────────────────────────────────────────────────────────────

def get_shares_outstanding_api(ticker: str) -> int:
    api_key = os.getenv("FMP_API_KEY")
    if not api_key:
        return 0
    r = requests.get(
        f"https://financialmodelingprep.com/api/v3/profile/{ticker}",
        params={"apikey": api_key},
        timeout=20,
    )
    if r.ok:
        data = r.json()
        if isinstance(data, list) and data:
            return int(data[0].get("sharesOutstanding", 0) or 0)
    return 0

# ────────────────────────────────────────────────────────────────────────────────
# Streamlit page config
# ────────────────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="DCF Analysis", page_icon="💵", layout="wide")

# ────────────────────────────────────────────────────────────────────────────────
# Helper functions to pivot FMP JSON → tidy DataFrames
# ────────────────────────────────────────────────────────────────────────────────
def _clean(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.set_index("date").drop(
        columns=[
            "symbol",
            "reportedCurrency",
            "cik",
            "fillingDate",
            "acceptedDate",
            "calendarYear",
            "period",
            "link",
            "finalLink",
        ],
        errors="ignore",
    )
    return df.T

def transform_income_statement(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df[df.columns[::-1]]  # newest→oldest flipped

def transform_balance_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df.reindex(sorted(df.columns, key=lambda d: pd.to_datetime(d)), axis=1)

def transform_cash_flow(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df.reindex(sorted(df.columns, key=lambda d: pd.to_datetime(d)), axis=1)

def pull_historical_data(ticker: str):
    inc = get_financial_statement(ticker, "income-statement")
    bs  = get_financial_statement(ticker, "balance-sheet-statement")
    cf  = get_financial_statement(ticker, "cash-flow-statement")
    return (
        transform_income_statement(inc),
        transform_balance_sheet(bs),
        transform_cash_flow(cf),
    )

# ────────────────────────────────────────────────────────────────────────────────
# Session state flags to control flow
# ────────────────────────────────────────────────────────────────────────────────
if "show_next" not in st.session_state:
    st.session_state.show_next = False
if "prev_ticker" not in st.session_state:
    st.session_state.prev_ticker = None

# ────────────────────────────────────────────────────────────────────────────────
# UI – ticker input only on first load
# ────────────────────────────────────────────────────────────────────────────────
st.markdown("<h1 style='text-align:center;'>DCF Analysis</h1>", unsafe_allow_html=True)
raw = st.text_input("Enter a stock ticker (e.g., AAPL, MSFT)")
ticker = raw.strip().upper() if raw else ""

# reset "next-stage" flag when ticker changes
if ticker != st.session_state.prev_ticker:
    st.session_state.show_next = False
    st.session_state.prev_ticker = ticker

# ────────────────────────────────────────────────────────────────────────────────
# Pull & display historicals once a ticker is entered
# ────────────────────────────────────────────────────────────────────────────────
if ticker:
    income_df, balance_df, cashflow_df = pull_historical_data(ticker)

    if income_df.empty:
        st.error(f"No financial-statement data returned for ticker: {ticker}")
        st.stop()

    st.markdown("### Historical Financial Statements")
    c1, c2, c3 = st.columns([1, 1, 1], gap="large")
    with c1:
        st.subheader("Income Statement")
        st.dataframe(income_df, height=300, use_container_width=True)
    with c2:
        st.subheader("Balance Sheet")
        st.dataframe(balance_df, height=300, use_container_width=True)
    with c3:
        st.subheader("Cash Flow Statement")
        st.dataframe(cashflow_df, height=300, use_container_width=True)

    st.markdown("---")
    c1, c2, c3, c4 = st.columns([1, 1, 1, 0.6], gap="medium")
    with c1:
        forecast_years = st.number_input(
            "Forecast Period (years)",
            min_value=1,
            max_value=10,
            value=5,
        )
    with c2:
        lt_growth = (
            st.number_input(
                "LT Growth Rate (%)",
                min_value=0.0,
                max_value=5.0,
                value=2.0,
                step=0.1,
                format="%.2f",
            )
            / 100
        )
    with c3:
        metrics = st.multiselect(
            "Variables to forecast",
            ["revenue", "gross_margin", "operating_margin", "tax_rate"],
            default=["revenue"],
        )
    with c4:
        if st.button("Next"):
            st.session_state.show_next = True

    # ────────────────────────────────────────────────────────────────────────────
    # Assumptions & Inputs – appears after "Next"
    # ────────────────────────────────────────────────────────────────────────────
    if st.session_state.show_next:
        st.markdown("### Assumptions & Inputs")
        a1, a2 = st.columns(2, gap="large")
        with a1:
            # Calculate default revenue growth from last two years
            last_year_revenue = income_df.loc["revenue"].iloc[-1]
            prev_year_revenue = income_df.loc["revenue"].iloc[-2]
            rev_growth_default = ((last_year_revenue / prev_year_revenue) - 1) * 100
            
            # Revenue growth is always shown since it's the base for all calculations
            rev_growth = (
                st.number_input(
                    "Revenue YoY Growth (%)",
                    min_value=-100.0,
                    max_value=200.0,
                    value=rev_growth_default,
                    step=1.0,
                    format="%.2f",
                )
                / 100
            )
            
            # Only show Gross Margin input if selected
            if "gross_margin" in metrics:
                gm_default = (
                    income_df.loc["grossProfit"].iloc[-1]
                    / income_df.loc["revenue"].iloc[-1]
                    * 100
                )
                gm = (
                    st.number_input(
                        "Gross Margin (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=gm_default,
                        step=0.1,
                        format="%.2f",
                    )
                    / 100
                )
            else:
                # Keep historical gross margin
                gm = income_df.loc["grossProfit"].iloc[-1] / income_df.loc["revenue"].iloc[-1]
            
            # Only show Operating Margin input if selected
            if "operating_margin" in metrics:
                om_default = (
                    income_df.loc["operatingIncome"].iloc[-1]
                    / income_df.loc["revenue"].iloc[-1]
                    * 100
                )
                om = (
                    st.number_input(
                        "Operating Margin (%)",
                        min_value=0.0,
                        max_value=100.0,
                        value=om_default,
                        step=0.1,
                        format="%.2f",
                    )
                    / 100
                )
            else:
                # Keep historical operating margin
                om = income_df.loc["operatingIncome"].iloc[-1] / income_df.loc["revenue"].iloc[-1]
            
            # Only show Tax Rate input if selected
            if "tax_rate" in metrics:
                tax_expense_default = (
                    income_df.loc["incomeTaxExpense"].iloc[-1]
                    / income_df.loc["revenue"].iloc[-1]
                    * 100
                )
                tax_rate = (
                    st.number_input(
                        "Tax Rate (%)",
                        min_value=0.0,
                        max_value=40.0,
                        value=tax_expense_default,
                        step=0.5,
                        format="%.2f",
                    )
                    / 100
                )
            else:
                # Keep historical tax rate
                tax_rate = income_df.loc["incomeTaxExpense"].iloc[-1] / income_df.loc["revenue"].iloc[-1]

        with a2:
            # D&A is always shown since it's needed for FCF calculation
            da_default = (
                income_df.loc["depreciationAndAmortization"].iloc[-1]
                / income_df.loc["revenue"].iloc[-1]
                * 100
            )
            da_pct = (
                st.number_input(
                    "D&A (% of Revenue)",
                    min_value=0.0,
                    max_value=100.0,
                    value=da_default,
                    step=0.1,
                    format="%.2f",
                )
                / 100
            )
            
            # CapEx is always shown since it's needed for FCF calculation
            capex_default = (
                cashflow_df.loc["capitalExpenditure"].iloc[-1]
                / income_df.loc["revenue"].iloc[-1]
                * 100
            )
            capex_pct = (
                st.number_input(
                    "CapEx (% of Revenue)",
                    min_value=0.0,
                    max_value=100.0,
                    value=abs(capex_default),
                    step=0.1,
                    format="%.2f",
                )
                / 100
            )
            
            # Other Income/Expense is always shown since it's needed for FCF calculation
            other_income_default = (
                income_df.loc["totalOtherIncomeExpensesNet"].iloc[-1]
                / income_df.loc["revenue"].iloc[-1]
                * 100
            )
            other_income_pct = (
                st.number_input(
                    "Other Income/Expense (% of Revenue)",
                    min_value=-100.0,
                    max_value=100.0,
                    value=other_income_default,
                    step=0.1,
                    format="%.2f",
                )
                / 100
            )
            
            # Interest Expense is always shown since it's needed for FCF calculation
            interest_expense_default = (
                income_df.loc["interestExpense"].iloc[-1]
                / income_df.loc["revenue"].iloc[-1]
                * 100
            )
            interest_expense_pct = (
                st.number_input(
                    "Interest Expense (% of Revenue)",
                    min_value=0.0,
                    max_value=100.0,
                    value=interest_expense_default,
                    step=0.1,
                    format="%.2f",
                )
                / 100
            )
            
            # Discount Rate is always shown since it's needed for DCF calculation
            discount_rate = (
                st.number_input(
                    "Discount Rate (%)",
                    min_value=1.0,
                    max_value=20.0,
                    value=10.0,
                    step=0.1,
                    format="%.2f",
                )
                / 100
            )

            st.markdown("---")
            if st.button("Generate DCF"):
                # ────────────────────────────────────────────────────────────────
                # Build workbook
                # ────────────────────────────────────────────────────────────────
                wb = Workbook()
                wb.remove(wb.active)

                # Historical tabs
                for df, name in zip(
                    (income_df, balance_df, cashflow_df),
                    ("Income Statement", "Balance Sheet", "Cash Flow Statement"),
                ):
                    ws = wb.create_sheet(title=name)
                    for row in dataframe_to_rows(df, index=True, header=True):
                        ws.append(row)

                # Assumptions tab
                aw = wb.create_sheet(title="Assumptions")
                aw.append(["Input", "Value"])
                for label, val in [
                    ("Revenue Growth (%)", rev_growth),
                    ("LT Growth Rate (%)", lt_growth),
                    ("Gross Margin (%)", gm),
                    ("Operating Margin (%)", om),
                    ("Tax Rate (%)", tax_rate),
                    ("D&A (% of Revenue)", da_pct),
                    ("CapEx (% of Revenue)", capex_pct),
                    ("Other Income/Expense (% of Revenue)", other_income_pct),
                    ("Interest Expense (% of Revenue)", interest_expense_pct),
                    ("Discount Rate (%)", discount_rate),
                ]:
                    aw.append([label, val])
                # Shares Outstanding
                idx_map_sh = {k: i + 2 for i, k in enumerate(income_df.index)}
                shares = (
                    income_df.iloc[idx_map_sh.get("weightedAverageShsOut", -1) - 2, -1]
                    if "weightedAverageShsOut" in idx_map_sh
                    else None
                )
                if not shares:
                    shares = get_shares_outstanding_api(ticker)
                aw.append(["Shares Outstanding", shares])

                # ────────────────────────────────────────────────────────────────
                # DCF Forecast tab
                # ────────────────────────────────────────────────────────────────
                fw = wb.create_sheet(title="DCF Forecast")

                hist = [str(d) for d in income_df.columns]
                last_dt = pd.to_datetime(income_df.columns[-1])
                fut = [
                    (last_dt + pd.DateOffset(years=i + 1)).strftime("%Y-%m-%d")
                    for i in range(forecast_years)
                ]
                fw.append(["Variable"] + hist + fut)

                # build rows (including your new items + the original FCF/EV rows)
                rows = [
                    "Revenue",
                    "COGS",
                    "Gross Profit",
                    "Gross Margin",
                    "Operating Costs",
                    "Operating Income",
                    "Operating Margin",
                    "interestExpense",
                    "totalOtherIncomeExpensesNet",
                    "Pre Tax Income",
                    "Income Tax",
                    "Net Income",
                    "Net Income Margin",
                    "Free Cash Flow",
                    "Free Cash Flow to Equity",
                    "Discount Factor",
                    "PV of FCF",
                    "Terminal Value",
                    "PV of Terminal Value",
                    "Enterprise Value",
                ]
                mapping = {}
                for r in rows:
                    fw.append([r] + [""] * (len(hist) + forecast_years))
                    mapping[r] = fw.max_row

                # dynamic row-map for Income Statement sheet
                idx_map_inc = {k: i + 2 for i, k in enumerate(income_df.index)}

                # historical links – your 5 yrs from Income Statement
                for i in range(len(hist)):
                    col = get_column_letter(i + 2)

                    fw[f"{col}{mapping['Revenue']}"]         = f"='Income Statement'!{col}3"
                    fw[f"{col}{mapping['COGS']}"]            = f"='Income Statement'!{col}4"
                    fw[f"{col}{mapping['Gross Profit']}"]    = f"='Income Statement'!{col}5"
                    fw[f"{col}{mapping['Gross Margin']}"]   = f"={col}{mapping['Gross Profit']}/{col}{mapping['Revenue']}"
                    fw[f"{col}{mapping['Operating Costs']}"] = f"='Income Statement'!{col}12"
                    fw[f"{col}{mapping['Operating Income']}"] = f"='Income Statement'!{col}19"
                    fw[f"{col}{mapping['Operating Margin']}"]   = f"={col}{mapping['Operating Income']}/{col}{mapping['Revenue']}"
                    fw[f"{col}{mapping['interestExpense']}"] = f"='Income Statement'!{col}15"
                    fw[f"{col}{mapping['totalOtherIncomeExpensesNet']}"] = f"='Income Statement'!{col}21"
                    # Pre Tax Income = OpInc – interest + otherIncome
                    fw[f"{col}{mapping['Pre Tax Income']}"]  = (
                        f"={col}{mapping['Operating Income']}-"
                        f"{col}{mapping['interestExpense']}+"
                        f"{col}{mapping['totalOtherIncomeExpensesNet']}"
                    )
                    # Income Tax & Net Income
                    fw[f"{col}{mapping['Income Tax']}"]      = f"={col}{mapping['Pre Tax Income']}*Assumptions!$B$6"
                    fw[f"{col}{mapping['Net Income']}"]      = f"={col}{mapping['Pre Tax Income']}-{col}{mapping['Income Tax']}"
                    fw[f"{col}{mapping['Net Income Margin']}"] = f"={col}{mapping['Net Income']}/{col}{mapping['Revenue']}"


                # forecast formulas for every period
                for i in range(forecast_years):
                    idx = len(hist) + i + 2
                    col = get_column_letter(idx)
                    prev = get_column_letter(idx - 1)

                    fw[f"{col}{mapping['Revenue']}"]         = f"={prev}{mapping['Revenue']}*(1+Assumptions!$B$2)"
                    fw[f"{col}{mapping['COGS']}"]            = f"={col}{mapping['Revenue']}*(1-Assumptions!$B$4)"
                    fw[f"{col}{mapping['Gross Profit']}"]    = f"={col}{mapping['Revenue']}-{col}{mapping['COGS']}"
                    fw[f"{col}{mapping['Gross Margin']}"]   = f"={col}{mapping['Gross Profit']}/{col}{mapping['Revenue']}"
                    fw[f"{col}{mapping['Operating Income']}"] = f"={col}{mapping['Revenue']}*Assumptions!$B$5"
                    fw[f"{col}{mapping['Operating Costs']}"] = f"={col}{mapping['Gross Profit']}-{col}{mapping['Operating Income']}"
                    fw[f"{col}{mapping['Operating Margin']}"]   = f"={col}{mapping['Operating Income']}/{col}{mapping['Revenue']}"

                    # project interest & other income based on revenue
                    fw[f"{col}{mapping['interestExpense']}"]            = f"={col}{mapping['Revenue']}*Assumptions!$B$10"
                    fw[f"{col}{mapping['totalOtherIncomeExpensesNet']}"] = f"={col}{mapping['Revenue']}*Assumptions!$B$9"

                    # Pre Tax, Tax, Net
                    fw[f"{col}{mapping['Pre Tax Income']}"]  = (
                        f"={col}{mapping['Operating Income']}-"
                        f"{col}{mapping['interestExpense']}+"
                        f"{col}{mapping['totalOtherIncomeExpensesNet']}"
                    )
                    fw[f"{col}{mapping['Income Tax']}"]      = f"={col}{mapping['Pre Tax Income']}*Assumptions!$B$6"
                    fw[f"{col}{mapping['Net Income']}"]      = f"={col}{mapping['Pre Tax Income']}-{col}{mapping['Income Tax']}"
                    fw[f"{col}{mapping['Net Income Margin']}"] = f"={col}{mapping['Net Income']}/{col}{mapping['Revenue']}"

                    # FCF, Equity & Discount
                    fw[f"{col}{mapping['Free Cash Flow']}"]            = (
                        f"={col}{mapping['Net Income']}+"
                        f"{col}{mapping['Revenue']}*Assumptions!$B$7-"
                        f"{col}{mapping['Revenue']}*Assumptions!$B$8"
                    )
                    fw[f"{col}{mapping['Free Cash Flow to Equity']}"] = f"={col}{mapping['Free Cash Flow']}"
                    fw[f"{col}{mapping['Discount Factor']}"]          = f"=1/(1+Assumptions!$B$11)^{i+1}"
                    fw[f"{col}{mapping['PV of FCF']}"]                = f"={col}{mapping['Free Cash Flow']}*{col}{mapping['Discount Factor']}"

                # Terminal + PV + EV + Price (unchanged)
                term_col = get_column_letter(len(hist) + forecast_years + 1)
                fw[f"{term_col}{mapping['Terminal Value']}"]       = (
                    f"={term_col}{mapping['Free Cash Flow to Equity']}*(1+Assumptions!$B$3)/"
                    f"(Assumptions!$B$11-Assumptions!$B$3)"
                )
                fw[f"{term_col}{mapping['PV of Terminal Value']}"] = (
                    f"={term_col}{mapping['Terminal Value']}*1/"
                    f"(1+Assumptions!$B$11)^{forecast_years}"
                )
                fw[f"B{mapping['Enterprise Value']}"]               = (
                    f"=SUM(C{mapping['PV of FCF']}:{term_col}{mapping['PV of FCF']})+"
                    f"{term_col}{mapping['PV of Terminal Value']}"
                )
                so_row = mapping["Enterprise Value"] + 1
                fw[f"A{so_row}"] = "Shares Outstanding"
                fw[f"B{so_row}"] = shares
                sp_row = so_row + 1
                fw[f"A{sp_row}"] = "Share Price"
                fw[f"B{sp_row}"] = f"=B{mapping['Enterprise Value']}/B{so_row}"

                # ────────────────────────────────────────────────────────────────
                # Send to user
                # ────────────────────────────────────────────────────────────────
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                st.download_button(
                    "📥 Download DCF Excel File",
                    data=buffer,
                    file_name=f"{ticker}_DCF.xlsx",
                    mime=(
                        "application/vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),
                )