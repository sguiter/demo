import os
import streamlit as st
import pandas as pd
import requests
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import openpyxl.styles
from streamlit_app.helpers.excel_formatting import apply_base_formatting, format_header
from streamlit_app.helpers.API_helpers import get_financial_statement

RATIO_ITEMS = {
    "grossProfitRatio", "ebitdaratio", "operatingIncomeRatio",
    "incomeBeforeTaxRatio", "netIncomeRatio", "eps", "epsdiluted",
}

ACRONYM_MAP = {
    # Income Statement
    "revenue": "Revenue",
    "costOfRevenue": "Cost of Revenue",
    "grossProfit": "Gross Profit",
    "grossProfitRatio": "Gross Profit %",
    "researchAndDevelopmentExpenses": "R&D Expense",
    "generalAndAdministrativeExpenses": "G&A Expense",
    "sellingAndMarketingExpenses": "S&M Expense",
    "sellingGeneralAndAdministrativeExpenses": "SG&A Expense",
    "otherExpenses": "Other Expenses",
    "operatingExpenses": "Operating Expenses",
    "costAndExpenses": "Cost & Expenses",
    "interestIncome": "Interest Income",
    "interestExpense": "Interest Expense",
    "depreciationAndAmortization": "Depreciation & Amortization",
    "ebitda": "EBITDA",
    "ebitdaratio": "EBITDA %",
    "operatingIncome": "Operating Income",
    "operatingIncomeRatio": "Operating Income %",
    "totalOtherIncomeExpensesNet": "Total Other Income & Expenses",
    "incomeBeforeTax": "Income Before Tax",
    "incomeBeforeTaxRatio": "Income Before Tax %",
    "incomeTaxExpense": "Income Tax Expense",
    "netIncome": "Net Income",
    "netIncomeRatio": "Net Income %",
    "eps": "EPS",
    "epsdiluted": "EPS Diluted",
    "weightedAverageShsOut": "Wtd Avg Shares Outstanding",
    "weightedAverageShsOutDil": "Wtd Avg Shares Outstanding Diluted",
    # Balance Sheet
    "cashAndCashEquivalents": "Cash & Cash Equivalents",
    "shortTermInvestments": "Short-Term Investments",
    "cashAndShortTermInvestments": "Cash & Short-Term Investments",
    "netReceivables": "Net Receivables",
    "inventory": "Inventory",
    "otherCurrentAssets": "Other Current Assets",
    "totalCurrentAssets": "Total Current Assets",
    "propertyPlantEquipmentNet": "PP&E, Net",
    "goodwill": "Goodwill",
    "intangibleAssets": "Intangible Assets",
    "goodwillAndIntangibleAssets": "Goodwill & Intangible Assets",
    "longTermInvestments": "Long-Term Investments",
    "taxAssets": "Tax Assets",
    "otherNonCurrentAssets": "Other Non-Current Assets",
    "totalNonCurrentAssets": "Total Non-Current Assets",
    "otherAssets": "Other Assets",
    "totalAssets": "Total Assets",
    "accountPayables": "Accounts Payable",
    "shortTermDebt": "Short-Term Debt",
    "taxPayables": "Tax Payables",
    "deferredRevenue": "Deferred Revenue",
    "otherCurrentLiabilities": "Other Current Liabilities",
    "totalCurrentLiabilities": "Total Current Liabilities",
    "longTermDebt": "Long-Term Debt",
    "deferredRevenueNonCurrent": "Deferred Revenue (Non-Current)",
    "deferredTaxLiabilitiesNonCurrent": "Deferred Tax Liabilities (Non-Current)",
    "otherNonCurrentLiabilities": "Other Non-Current Liabilities",
    "totalNonCurrentLiabilities": "Total Non-Current Liabilities",
    "otherLiabilities": "Other Liabilities",
    "capitalLeaseObligations": "Capital Lease Obligations",
    "totalLiabilities": "Total Liabilities",
    "preferredStock": "Preferred Stock",
    "commonStock": "Common Stock",
    "retainedEarnings": "Retained Earnings",
    "accumulatedOtherComprehensiveIncomeLoss": "Accumulated Other Comprehensive Income (Loss)",
    "othertotalStockholdersEquity": "Other Total Stockholders' Equity",
    "totalStockholdersEquity": "Total Stockholders' Equity",
    "totalEquity": "Total Equity",
    "totalLiabilitiesAndStockholdersEquity": "Total Liabilities & Stockholders' Equity",
    "minorityInterest": "Minority Interest",
    "totalLiabilitiesAndTotalEquity": "Total Liabilities & Total Equity",
    "totalInvestments": "Total Investments",
    "totalDebt": "Total Debt",
    "netDebt": "Net Debt",
    # Cash Flow
    "deferredIncomeTax": "Deferred Income Tax",
    "stockBasedCompensation": "SBC",
    "changeInWorkingCapital": "Change in Working Capital",
    "accountsReceivables": "Accounts Receivable",
    "accountsPayables": "Accounts Payable",
    "otherWorkingCapital": "Other Working Capital",
    "otherNonCashItems": "Other Non-Cash Items",
    "netCashProvidedByOperatingActivities": "Cash Flow from Operations (CFO)",
    "investmentsInPropertyPlantAndEquipment": "Investments in PP&E",
    "acquisitionsNet": "Acquisitions, Net",
    "purchasesOfInvestments": "Purchases of Investments",
    "salesMaturitiesOfInvestments": "Sales/Maturities of Investments",
    "otherInvestingActivites": "Other Investing Activities",
    "netCashUsedForInvestingActivites": "Cash Flow Used in Investing (CFI)",
    "debtRepayment": "Debt Repayment",
    "commonStockIssued": "Common Stock Issued",
    "commonStockRepurchased": "Common Stock Repurchased",
    "dividendsPaid": "Dividends Paid",
    "otherFinancingActivites": "Other Financing Activities",
    "netCashUsedProvidedByFinancingActivities": "Cash Flow from Financing (CFF)",
    "effectOfForexChangesOnCash": "Effect of Forex Changes on Cash",
    "netChangeInCash": "Net Change in Cash",
    "cashAtEndOfPeriod": "Cash at End of Period",
    "cashAtBeginningOfPeriod": "Cash at Beginning of Period",
    "operatingCashFlow": "Operating Cash Flow",
    "capitalExpenditure": "CapEx",
    "freeCashFlow": "Free Cash Flow"
}

def get_shares_outstanding_api(ticker: str) -> int:
    api_key = os.getenv("FMP_API_KEY")
    if not api_key:
        return 0
    r = requests.get(
        f"https://financialmodelingprep.com/api/v3/profile/{ticker}",
        params={"apikey": api_key},
        timeout=20,
    )
    if r.ok:
        data = r.json()
        if isinstance(data, list) and data:
            return int(data[0].get("sharesOutstanding", 0) or 0)
    return 0

def autofit_columns(ws):
    for column_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            if cell.value is not None:
                length = len(str(cell.value))
                if length > max_length:
                    max_length = length
        ws.column_dimensions[col_letter].width = max_length + 2

# Streamlit page config
st.set_page_config(page_title="DCF Analysis", layout="wide")

def _clean(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    df = df.set_index("date").drop(
        columns=[
            "symbol","reportedCurrency","cik","fillingDate","acceptedDate",
            "calendarYear","period","link","finalLink"
        ],
        errors="ignore"
    )
    return df.T

def transform_income_statement(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df[df.columns[::-1]]

def transform_balance_sheet(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df.reindex(sorted(df.columns, key=lambda d: pd.to_datetime(d)), axis=1)

def transform_cash_flow(df: pd.DataFrame) -> pd.DataFrame:
    df = _clean(df)
    return df.reindex(sorted(df.columns, key=lambda d: pd.to_datetime(d)), axis=1)

def pull_historical_data(ticker: str):
    inc = get_financial_statement(ticker, "income-statement")
    bs  = get_financial_statement(ticker, "balance-sheet-statement")
    cf  = get_financial_statement(ticker, "cash-flow-statement")
    return (
        transform_income_statement(inc),
        transform_balance_sheet(bs),
        transform_cash_flow(cf),
    )

# Session state flags
if "show_next" not in st.session_state:
    st.session_state.show_next = False
if "prev_ticker" not in st.session_state:
    st.session_state.prev_ticker = None

# UI – ticker input
st.markdown("<h1 style='text-align:center;'>DCF Analysis</h1>", unsafe_allow_html=True)
raw = st.text_input("Enter a stock ticker (e.g., AAPL, MSFT)")
ticker = raw.strip().upper() if raw else ""
if ticker != st.session_state.prev_ticker:
    st.session_state.show_next = False
    st.session_state.prev_ticker = ticker

if ticker:
    income_df, balance_df, cashflow_df = pull_historical_data(ticker)
    if income_df.empty:
        st.error(f"No financial-statement data returned for ticker: {ticker}")
        st.stop()

    # Historical statements
    st.markdown("### Historical Financial Statements")
    c1, c2, c3 = st.columns([1,1,1], gap="large")
    with c1:
        st.subheader("Income Statement")
        st.dataframe(income_df, height=300, use_container_width=True)
    with c2:
        st.subheader("Balance Sheet")
        st.dataframe(balance_df, height=300, use_container_width=True)
    with c3:
        st.subheader("Cash Flow Statement")
        st.dataframe(cashflow_df, height=300, use_container_width=True)

    st.markdown("---")
    c1, c2, c3, c4 = st.columns([1,1,1,0.6], gap="medium")
    with c1:
        forecast_years = st.number_input("Forecast Period (years)", min_value=1, max_value=10, value=5)
    with c2:
        lt_growth = st.number_input("LT Growth Rate (%)", min_value=0.0, max_value=5.0, value=2.0, step=0.1, format="%.2f")/100
    with c3:
        metrics = st.multiselect("Variables to forecast", ["revenue","gross_margin","operating_margin","tax_rate"], default=["revenue"])
    with c4:
        if st.button("Next"):
            st.session_state.show_next = True

    if st.session_state.show_next:
        # Assumptions & Inputs
        st.markdown("### Assumptions & Inputs")
        a1, a2 = st.columns(2, gap="large")
        with a1:
            last_rev = income_df.loc["revenue"].iloc[-1]
            prev_rev = income_df.loc["revenue"].iloc[-2]
            rev_growth_default = ((last_rev/prev_rev) - 1) * 100
            rev_growth = st.number_input("Revenue YoY Growth (%)", min_value=-100.0, max_value=200.0, value=rev_growth_default, step=1.0, format="%.2f")/100

            if "gross_margin" in metrics:
                gm_default = income_df.loc["grossProfit"].iloc[-1] / last_rev * 100
                gm = st.number_input("Gross Margin (%)", min_value=0.0, max_value=100.0, value=gm_default, step=0.1, format="%.2f")/100
            else:
                gm = income_df.loc["grossProfit"].iloc[-1] / last_rev

            if "operating_margin" in metrics:
                om_default = income_df.loc["operatingIncome"].iloc[-1] / last_rev * 100
                om = st.number_input("Operating Margin (%)", min_value=0.0, max_value=100.0, value=om_default, step=0.1, format="%.2f")/100
            else:
                om = income_df.loc["operatingIncome"].iloc[-1] / last_rev

            if "tax_rate" in metrics:
                tx_def = income_df.loc["incomeTaxExpense"].iloc[-1] / last_rev * 100
                tax_rate = st.number_input("Tax Rate (%)", min_value=0.0, max_value=40.0, value=tx_def, step=0.5, format="%.2f")/100
            else:
                tax_rate = income_df.loc["incomeTaxExpense"].iloc[-1] / last_rev

        with a2:
            da_def = income_df.loc["depreciationAndAmortization"].iloc[-1] / last_rev * 100
            da_pct = st.number_input("D&A (% of Revenue)", min_value=0.0, max_value=100.0, value=da_def, step=0.1, format="%.2f")/100

            cap_def = cashflow_df.loc["capitalExpenditure"].iloc[-1] / last_rev * 100
            capex_pct = st.number_input("CapEx (% of Revenue)", min_value=0.0, max_value=100.0, value=abs(cap_def), step=0.1, format="%.2f")/100

            oi_def = income_df.loc["totalOtherIncomeExpensesNet"].iloc[-1] / last_rev * 100
            other_income_pct = st.number_input("Other Income/Expense (% of Revenue)", min_value=-100.0, max_value=100.0, value=oi_def, step=0.1, format="%.2f")/100

            ie_def = income_df.loc["interestExpense"].iloc[-1] / last_rev * 100
            interest_expense_pct = st.number_input("Interest Expense (% of Revenue)", min_value=0.0, max_value=100.0, value=ie_def, step=0.1, format="%.2f")/100

            discount_rate = st.number_input("Discount Rate (%)", min_value=1.0, max_value=20.0, value=10.0, step=0.1, format="%.2f")/100

        st.markdown("---")
        if st.button("Generate DCF"):
            wb = Workbook()
            wb.remove(wb.active)

            def add_statement_header(ws, cols, name):
                ws.sheet_view.showGridLines = False
                ws.append([])
                br = ws.max_row
                for c in range(1, cols+1):
                    ws.cell(row=br, column=c).border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(style='thin', color='000000'),
                    )
                format_header(ws, br+1, 1, cols, name)
                ws.append([])
                br = ws.max_row
                for c in range(1, cols+1):
                    ws.cell(row=br, column=c).border = openpyxl.styles.Border(
                        top=openpyxl.styles.Side(style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(style='thin', color='000000'),
                    )

            def add_statement_with_values(ws, df, name):
                ws.sheet_view.showGridLines = False
                start_row = ws.max_row + 4
                cols = len(df.columns) + 1
                add_statement_header(ws, cols, name)
                ws.append([""] + list(df.columns))
                for idx_name, row_data in df.iterrows():
                    display_name = ACRONYM_MAP.get(idx_name, idx_name)
                    row_values = [display_name] + list(row_data)
                    if idx_name not in RATIO_ITEMS:
                        row_values = [row_values[0]] + [
                            val/1_000_000 if isinstance(val, (int, float)) else val
                            for val in row_values[1:]
                        ]
                    ws.append(row_values)
                return start_row

            # ── Historicals ──
            hist_ws = wb.create_sheet("Historicals")
            hist_ws.sheet_view.showGridLines = False
            apply_base_formatting(hist_ws)

            # Income Statement
            add_statement_with_values(hist_ws, income_df, "Income Statement")
            # Separator line after Income Statement
            sep_row = hist_ws.max_row + 1
            for col in range(1, len(income_df.columns) + 2):
                hist_ws.cell(row=sep_row, column=col).border = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(style='thin', color='000000'),
                    bottom=openpyxl.styles.Side(style='thin', color='000000'),
                )

            # Balance Sheet
            add_statement_with_values(hist_ws, balance_df, "Balance Sheet")
            # Separator line after Balance Sheet
            sep_row = hist_ws.max_row + 1
            for col in range(1, len(balance_df.columns) + 2):
                hist_ws.cell(row=sep_row, column=col).border = openpyxl.styles.Border(
                    top=openpyxl.styles.Side(style='thin', color='000000'),
                    bottom=openpyxl.styles.Side(style='thin', color='000000'),
                )

            # Cash Flow Statement
            add_statement_with_values(hist_ws, cashflow_df, "Cash Flow Statement")

            # Finally, autofit all columns in the Historicals sheet
            autofit_columns(hist_ws)
            hist_ws.insert_rows(1)
            hist_ws["A1"] = "Note: All numbers are in millions"
            hist_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=hist_ws.max_column)
            hist_ws["A1"].font = Font(italic=True)

            # ── Assumptions ──
            aw = wb.create_sheet("Assumptions")
            aw.sheet_view.showGridLines = False
            aw.append(["Input", "Value"])
            aw.cell(1,1).font = Font(bold=True)
            aw.cell(1,2).font = Font(bold=True)
            assumptions_data = [
                ("Revenue Growth (%)", rev_growth),
                ("LT Growth Rate (%)", lt_growth),
                ("Gross Margin (%)", gm),
                ("Operating Margin (%)", om),
                ("Tax Rate (%)", tax_rate),
                ("D&A (% of Revenue)", da_pct),
                ("CapEx (% of Revenue)", capex_pct),
                ("Other Income/Expense (% of Revenue)", other_income_pct),
                ("Interest Expense (% of Revenue)", interest_expense_pct),
                ("Discount Rate (%)", discount_rate),
            ]
            row = 2
            for label, val in assumptions_data:
                aw.append([label, val])
                aw.cell(row,2).number_format = '0.00%'
                row += 1
            idx_map_sh = {k: i+2 for i,k in enumerate(income_df.index)}
            shares = (
                income_df.iloc[idx_map_sh.get("weightedAverageShsOut",-1)-2,-1]
                if "weightedAverageShsOut" in idx_map_sh else get_shares_outstanding_api(ticker)
            )
            aw.append(["Shares Outstanding", shares])
            aw.cell(row,2).number_format = '#,##0'
            aw.column_dimensions['A'].width = 35
            aw.column_dimensions['B'].width = 15
            # ── Add million-scaling note to Assumptions ──
            aw.insert_rows(1)
            aw["A1"] = "Note: All numbers are in millions"
            aw.merge_cells(start_row=1, start_column=1, end_row=1, end_column=aw.max_column)
            aw["A1"].font = Font(italic=True)

            # ── DCF Forecast ──
            fw = wb.create_sheet("DCF Forecast")
            fw.sheet_view.showGridLines = False
            hist = [str(d) for d in income_df.columns]
            last_dt = pd.to_datetime(income_df.columns[-1])
            fut = [(last_dt + pd.DateOffset(years=i+1)).strftime("%Y-%m-%d") for i in range(forecast_years)]
            fw.append(["Variable"] + hist + fut)

            rows = [
                "revenue","costOfRevenue","grossProfit","grossProfitRatio",
                "operatingExpenses","operatingIncome","operatingIncomeRatio",
                "interestExpense","totalOtherIncomeExpensesNet","Pre Tax Income",
                "incomeTaxExpense","netIncome","netIncomeRatio",
                "Free Cash Flow","Free Cash Flow to Equity",
                "Discount Factor","PV of FCF","Terminal Value",
                "PV of Terminal Value","Enterprise Value",
            ]
            mapping = {}
            for r in rows:
                fw.append([ACRONYM_MAP.get(r,r)] + [""]*(len(hist)+forecast_years))
                mapping[r] = fw.max_row

            forecast_start_col = len(hist)+2
            forecast_end_col = forecast_start_col + forecast_years - 1

            # bold forecast headers & values
            for c in range(forecast_start_col, forecast_end_col+1):
                fw.cell(1,c).font = Font(color="000080", bold=False)

            # historical links
            inv = {v:k for k,v in ACRONYM_MAP.items()}
            needed = ["revenue","costOfRevenue","grossProfit","operatingExpenses",
                      "operatingIncome","interestExpense","totalOtherIncomeExpensesNet",
                      "incomeTaxExpense"]
            hist_rows = {}
            for r in range(1, hist_ws.max_row+1):
                val = hist_ws.cell(row=r, column=1).value
                raw = inv.get(val)
                if raw in needed:
                    hist_rows[raw] = r

            for i in range(len(hist)):
                col = get_column_letter(i+2)
                fw[f"{col}{mapping['revenue']}"] = f"=Historicals!{col}{hist_rows['revenue']}"
                fw[f"{col}{mapping['costOfRevenue']}"] = f"=Historicals!{col}{hist_rows['costOfRevenue']}"
                fw[f"{col}{mapping['grossProfit']}"] = f"=Historicals!{col}{hist_rows['grossProfit']}"
                fw[f"{col}{mapping['grossProfitRatio']}"] = f"={col}{mapping['grossProfit']}/{col}{mapping['revenue']}"
                fw[f"{col}{mapping['operatingExpenses']}"] = f"=Historicals!{col}{hist_rows['operatingExpenses']}"
                fw[f"{col}{mapping['operatingIncome']}"] = f"=Historicals!{col}{hist_rows['operatingIncome']}"
                fw[f"{col}{mapping['operatingIncomeRatio']}"] = f"={col}{mapping['operatingIncome']}/{col}{mapping['revenue']}"
                fw[f"{col}{mapping['interestExpense']}"] = f"=Historicals!{col}{hist_rows['interestExpense']}"
                fw[f"{col}{mapping['totalOtherIncomeExpensesNet']}"] = f"=Historicals!{col}{hist_rows['totalOtherIncomeExpensesNet']}"
                fw[f"{col}{mapping['Pre Tax Income']}"] = (
                    f"={col}{mapping['operatingIncome']}-"
                    f"{col}{mapping['interestExpense']}+"
                    f"{col}{mapping['totalOtherIncomeExpensesNet']}"
                )
                fw[f"{col}{mapping['incomeTaxExpense']}"] = f"=Historicals!{col}{hist_rows['incomeTaxExpense']}"
                fw[f"{col}{mapping['netIncome']}"] = f"={col}{mapping['Pre Tax Income']}-{col}{mapping['incomeTaxExpense']}"
                fw[f"{col}{mapping['netIncomeRatio']}"] = f"={col}{mapping['netIncome']}/{col}{mapping['revenue']}"

            # forecast formulas
            for i in range(forecast_years):
                idx = len(hist) + i + 2
                col = get_column_letter(idx)
                prev = get_column_letter(idx-1)
                fw[f"{col}{mapping['revenue']}"] = f"={prev}{mapping['revenue']}*(1+Assumptions!$B$3)"
                fw[f"{col}{mapping['costOfRevenue']}"] = f"={col}{mapping['revenue']}*(1-Assumptions!$B$5)"
                fw[f"{col}{mapping['grossProfit']}"] = f"={col}{mapping['revenue']}-{col}{mapping['costOfRevenue']}"
                fw[f"{col}{mapping['grossProfitRatio']}"] = f"={col}{mapping['grossProfit']}/{col}{mapping['revenue']}"
                fw[f"{col}{mapping['operatingExpenses']}"] = f"={col}{mapping['grossProfit']}-{col}{mapping['operatingIncome']}"
                fw[f"{col}{mapping['operatingIncome']}"] = f"={col}{mapping['revenue']}*Assumptions!$B$6"
                fw[f"{col}{mapping['operatingIncomeRatio']}"] = f"={col}{mapping['operatingIncome']}/{col}{mapping['revenue']}"
                fw[f"{col}{mapping['interestExpense']}"] = f"={col}{mapping['revenue']}*Assumptions!$B$11"
                fw[f"{col}{mapping['totalOtherIncomeExpensesNet']}"] = f"={col}{mapping['revenue']}*Assumptions!$B$10"
                fw[f"{col}{mapping['Pre Tax Income']}"] = (
                    f"={col}{mapping['operatingIncome']}-"
                    f"{col}{mapping['interestExpense']}+"
                    f"{col}{mapping['totalOtherIncomeExpensesNet']}"
                )
                fw[f"{col}{mapping['incomeTaxExpense']}"] = f"={col}{mapping['Pre Tax Income']}*Assumptions!$B$7"
                fw[f"{col}{mapping['netIncome']}"] = f"={col}{mapping['Pre Tax Income']}-{col}{mapping['incomeTaxExpense']}"
                fw[f"{col}{mapping['netIncomeRatio']}"] = f"={col}{mapping['netIncome']}/{col}{mapping['revenue']}"
                fw[f"{col}{mapping['Free Cash Flow']}"] = (
                    f"={col}{mapping['netIncome']}+"
                    f"{col}{mapping['revenue']}*Assumptions!$B$8-"
                    f"{col}{mapping['revenue']}*Assumptions!$B$9"
                )
                fw[f"{col}{mapping['Free Cash Flow to Equity']}"] = f"={col}{mapping['Free Cash Flow']}"
                fw[f"{col}{mapping['Discount Factor']}"] = f"=1/(1+Assumptions!$B$12)^{i+1}"
                fw[f"{col}{mapping['PV of FCF']}"] = f"={col}{mapping['Free Cash Flow']}*{col}{mapping['Discount Factor']}"

                # ── format margin rows as percentage ──
                for margin in ["grossProfitRatio", "operatingIncomeRatio", "netIncomeRatio"]:
                    row_idx = mapping[margin]
                    for col_idx in range(forecast_start_col, forecast_end_col + 1):
                        fw.cell(row=row_idx, column=col_idx).number_format = '0.00%'

            
                # ── Integer format all forecasted values (no decimals) ──
                for name, row_idx in mapping.items():
                    if name not in ["grossProfitRatio", "operatingIncomeRatio", "netIncomeRatio"]:
                        for c in range(forecast_start_col, forecast_end_col + 1):
                            fw.cell(row=row_idx, column=c).number_format = '#,##0'
                for name, row_idx in mapping.items():
                    for c in range(forecast_start_col, forecast_end_col + 1):
                        fw.cell(row=row_idx, column=c).font = Font(color="000080")    # navy color

            # Terminal & PV & EV & Price
            term_col = get_column_letter(len(hist) + forecast_years + 1)
            fw[f"{term_col}{mapping['Terminal Value']}"] = (
                f"={term_col}{mapping['Free Cash Flow to Equity']}*(1+Assumptions!$B$3)/"
                f"(Assumptions!$B$12-Assumptions!$B$3)"
            )
            fw[f"{term_col}{mapping['PV of Terminal Value']}"] = (
                f"={term_col}{mapping['Terminal Value']}*1/"
                f"(1+Assumptions!$B$12)^{forecast_years}"
            )
            fw[f"B{mapping['Enterprise Value']}"] = (
                f"=SUM(C{mapping['PV of FCF']}:{term_col}{mapping['PV of FCF']})+"
                f"{term_col}{mapping['PV of Terminal Value']}"
            )

            # Shares Outstanding
            so_row = mapping["Enterprise Value"] + 1
            fw[f"A{so_row}"] = "Shares Outstanding"
            fw[f"B{so_row}"] = "=Assumptions!B13"

            # Share Price – fix million‐scaling and format
            sp_row = so_row + 1
            fw[f"A{sp_row}"] = "Share Price"
            fw[f"B{sp_row}"] = f"=B{mapping['Enterprise Value']}*1000000/B{so_row}"
            sp_cell = fw.cell(row=sp_row, column=2)
            sp_cell.number_format = '$#,##0.00'    
            sp_cell.font = Font(bold=True)

            # ── format all margin rows (%) across entire DCF Forecast sheet ──
            for margin in ["grossProfitRatio", "operatingIncomeRatio", "netIncomeRatio"]:
                row_idx = mapping[margin]
                for col_idx in range(2, fw.max_column + 1):
                    fw.cell(row=row_idx, column=col_idx).number_format = '0.00%'
            
            # ── set uniform width for all DCF Forecast columns ──
            uniform_width = 15
            for col_idx in range(1, fw.max_column + 1):
                fw.column_dimensions[get_column_letter(col_idx)].width = uniform_width

            # Download
            buffer = BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            st.download_button(
                "📥 Download DCF Excel File",
                data=buffer,
                file_name=f"{ticker}_DCF.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
