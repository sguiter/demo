import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side, Alignment

from streamlit_app.helpers.API_helpers import get_financial_statement
from streamlit_app.helpers.base_formatting import formatting_page

# -------------------------
# Constants
# -------------------------

RATIO_ITEMS = {
    "grossProfitRatio",
    "ebitdaratio",
    "operatingIncomeRatio",
    "incomeBeforeTaxRatio",
    "netIncomeRatio",
    "eps",
    "epsdiluted"
}

SEPARATOR_BORDER = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="thin", color="000000")
)

# -------------------------
# Helper Functions
# -------------------------

def transform_income_statement(df):
    df = df.set_index('date').drop(columns=[
        'symbol','reportedCurrency','cik','fillingDate','acceptedDate',
        'calendarYear','period','link','finalLink'
    ], errors='ignore').transpose()
    return df[df.columns[::-1]]

def transform_balance_sheet(df):
    df = df.set_index('date').drop(columns=[
        'symbol','reportedCurrency','cik','fillingDate','acceptedDate',
        'calendarYear','period','link','finalLink'
    ], errors='ignore').transpose()
    return df[df.columns[::-1]]

def transform_cash_flow(df):
    df = df.set_index('date').drop(columns=[
        'symbol','reportedCurrency','cik','fillingDate','acceptedDate',
        'calendarYear','period','link','finalLink'
    ], errors='ignore').transpose()
    return df[df.columns[::-1]]

def pull_historical_data(ticker):
    inc = get_financial_statement(ticker, "income-statement")
    bal = get_financial_statement(ticker, "balance-sheet-statement")
    cf  = get_financial_statement(ticker, "cash-flow-statement")
    return (
        transform_income_statement(inc),
        transform_balance_sheet(bal),
        transform_cash_flow(cf)
    )

def create_historical_combined_sheet(wb, dfs, titles):
    ws = wb.create_sheet("Historical")
    start_row = 1
    for df, title in zip(dfs, titles):
        # Title + separator
        ws.cell(row=start_row, column=1, value=title)
        for col_idx in range(1, df.shape[1] + 2):
            ws.cell(row=start_row, column=col_idx).border = SEPARATOR_BORDER

        # Data rows (bold ratios)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=True, header=True), start=start_row+1):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                if r_idx > start_row+1 and row[0] in RATIO_ITEMS:
                    cell.font = Font(bold=True, name=cell.font.name, size=cell.font.size)
        start_row += df.shape[0] + 1 + 5

# -------------------------
# Streamlit UI
# -------------------------

st.title("💸 Discounted Cash Flow (DCF) Calculator")

ticker = st.text_input("Enter a stock ticker (e.g., AAPL, MSFT)")
forecast_years = st.slider("How many years to forecast?", 1, 10, 5)

st.subheader("Financial Assumptions")
revenue_growth          = st.number_input("Revenue Growth Rate (%)", 8.0)    / 100
gross_margin            = st.number_input("Gross Margin (%)", 60.0)         / 100
operating_costs_percent = st.number_input("Operating Costs (% of Revenue)", 30.0) / 100
tax_rate                = st.number_input("Tax Rate (%)", 21.0)             / 100
da_percent              = st.number_input("D&A (% of Revenue)", 3.0)        / 100
capex_percent           = st.number_input("CapEx (% of Revenue)", 5.0)      / 100

if st.button("Build DCF Model"):
    if not ticker:
        st.warning("Please enter a valid ticker.")
    else:
        with st.spinner("Building your DCF model..."):
            inc_df, bal_df, cf_df = pull_historical_data(ticker)
            inc_df.columns = pd.to_datetime(inc_df.columns).date
            bal_df.columns = pd.to_datetime(bal_df.columns).date
            cf_df.columns  = pd.to_datetime(cf_df.columns).date

            # Convert non-ratio items to millions
            for df in (inc_df, bal_df, cf_df):
                mask = [idx not in RATIO_ITEMS for idx in df.index]
                df.loc[mask] = df.loc[mask] / 1_000_000

            # Historical sheet
            wb = Workbook()
            wb.remove(wb.active)
            create_historical_combined_sheet(
                wb,
                dfs=[inc_df, bal_df, cf_df],
                titles=["Income Statement", "Balance Sheet", "Statement of Cash Flows"]
            )

            # Assumptions sheet
            asum = wb.create_sheet("Assumptions")
            asum.append(["Input", "Value"])
            asum.append(["Revenue Growth Rate (%)", revenue_growth])
            asum.append(["Gross Margin (%)", gross_margin])
            asum.append(["Operating Costs (% of Revenue)", operating_costs_percent])
            asum.append(["Tax Rate (%)", tax_rate])
            asum.append(["D&A (% of Revenue)", da_percent])
            asum.append(["CapEx (% of Revenue)", capex_percent])

            # DCF Forecast sheet
            forecast_ws = wb.create_sheet("DCF Forecast")
            lines = [
                "Revenue", "COGS", "Gross Profit", "Gross Margin %",
                "Operating Costs", "Operating Income", "Operating Margin %",
                "Income Tax", "Net Income", "Net Income Margin %", "Free Cash Flow"
            ]
            # Headers: historical dates + next years
            historical_headers = inc_df.columns.tolist()
            last_year = historical_headers[-1].year
            forecast_headers = [last_year + i + 1 for i in range(forecast_years)]
            years = historical_headers + forecast_headers
            forecast_ws.append(["Variable"] + years)

            # Bold the forecast year header cells
            num_hist = len(historical_headers)
            for col_idx in range(2 + num_hist, 2 + num_hist + forecast_years):
                cell = forecast_ws.cell(row=1, column=col_idx)
                cell.font = Font(bold=True, name=cell.font.name, size=cell.font.size)

            # Build rows and formulas
            row_map = {}
            for line in lines:
                forecast_ws.append([line] + [""] * len(years))
                row_map[line] = forecast_ws.max_row

            # Populate historical formulas
            for col_idx in range(2, 2 + len(inc_df.columns)):
                col = get_column_letter(col_idx)
                forecast_ws[f"{col}{row_map['Revenue']}"] = f"='Historical'!{col}4"
                lookup = {
                    "COGS":"costOfRevenue", "Gross Profit":"grossProfit",
                    "Operating Income":"operatingIncome", "Income Tax":"incomeTaxExpense",
                    "Net Income":"netIncome"
                }
                for var, key in lookup.items():
                    r = next(r for r,v in enumerate(inc_df.index, start=3) if v==key)
                    forecast_ws[f"{col}{row_map[var]}"] = f"='Historical'!{col}{r}"
                forecast_ws[f"{col}{row_map['Gross Margin %']}"]      = f"={col}{row_map['Gross Profit']}/{col}{row_map['Revenue']}"
                forecast_ws[f"{col}{row_map['Operating Margin %']}"]  = f"={col}{row_map['Operating Income']}/{col}{row_map['Revenue']}"
                forecast_ws[f"{col}{row_map['Net Income Margin %']}"] = f"={col}{row_map['Net Income']}/{col}{row_map['Revenue']}"

            # Populate forecast formulas
            for i in range(forecast_years):
                idx = 2 + len(inc_df.columns) + i
                col = get_column_letter(idx)
                prev = get_column_letter(idx - 1)
                forecast_ws[f"{col}{row_map['Revenue']}"]         = f"={prev}{row_map['Revenue']}*(1+Assumptions!B2)"
                forecast_ws[f"{col}{row_map['COGS']}"]            = f"={col}{row_map['Revenue']}*(1-Assumptions!B3)"
                forecast_ws[f"{col}{row_map['Gross Profit']}"]    = f"={col}{row_map['Revenue']}-{col}{row_map['COGS']}"
                forecast_ws[f"{col}{row_map['Operating Costs']}"] = f"={col}{row_map['Revenue']}*Assumptions!B4"
                forecast_ws[f"{col}{row_map['Operating Income']}"] = f"={col}{row_map['Gross Profit']}-{col}{row_map['Operating Costs']}"
                forecast_ws[f"{col}{row_map['Income Tax']}"]      = f"={col}{row_map['Operating Income']}*Assumptions!B5"
                forecast_ws[f"{col}{row_map['Net Income']}"]      = f"={col}{row_map['Operating Income']}-{col}{row_map['Income Tax']}"
                forecast_ws[f"{col}{row_map['Free Cash Flow']}"]  = (
                    f"={col}{row_map['Net Income']}+"
                    f"(Assumptions!B6*{col}{row_map['Revenue']})-"
                    f"(Assumptions!B7*{col}{row_map['Revenue']})"
                )

            # Formatting + dynamic sizing + left align
            for sheet in wb.worksheets:
                formatting_page(sheet)
                
                # Bold all forecasted years in DCF Forecast sheet
                if sheet.title == "DCF Forecast":
                    num_hist = len(historical_headers)
                    for row in range(1, sheet.max_row + 1):
                        for col in range(2 + num_hist, 2 + num_hist + forecast_years):
                            cell = sheet.cell(row=row, column=col)
                            cell.font = Font(bold=True, name=cell.font.name, size=cell.font.size)
                
                # Auto-size each column based on its content (after all formatting)
                for ci in range(1, sheet.max_column + 1):
                    letter = get_column_letter(ci)
                    col_width = max(
                        (len(str(cell.value).rstrip()) for cell in sheet[letter] if cell.value),
                        default=0
                    )
                    sheet.column_dimensions[letter].width = col_width + 1
                    for cell in sheet[letter]:
                        cell.alignment = Alignment(horizontal="left", vertical=cell.alignment.vertical)

            # Set a uniform column width for every column in the DCF Forecast worksheet
            uniform_width = 15  # You can adjust this value as needed
            for sheet in wb.worksheets:
                if sheet.title == "DCF Forecast":
                    for ci in range(1, sheet.max_column + 1):
                        letter = get_column_letter(ci)
                        sheet.column_dimensions[letter].width = uniform_width

            # Output
            buf = BytesIO()
            wb.save(buf)
            buf.seek(0)
            st.success(f"DCF Model for {ticker.upper()} built successfully!")
            st.download_button(
                "📥 Download DCF Excel File",
                data=buf,
                file_name=f"{ticker.upper()}_DCF_Model.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
